# coding: utf-8
# !/usr/bin/python

# This file is part of the FlOpEDT/FlOpScheduler project.
# Copyright (c) 2017
# Authors: Iulian Ober, Paul Renaud-Goud, Pablo Seban, et al.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public
# License along with this program. If not, see
# <http://www.gnu.org/licenses/>.
#
# You can be released from the requirements of the license by purchasing
# a commercial license. Buying such a license is mandatory as soon as
# you develop activities involving the FlOpEDT/FlOpScheduler software
# without disclosing the source code of your own applications.

import datetime as dt
import os

from django.conf import settings as ds
from django.db import transaction
from openpyxl import load_workbook

from base.models import (
    Course,
    CourseAdditional,
    CoursePossibleTutors,
    CourseType,
    Dependency,
    GenericGroup,
    Module,
    ModuleTutorRepartition,
    RoomType,
    SchedulingPeriod,
    TrainingPeriod,
)
from misc.assign_colors import assign_module_color
from people.models import Tutor
from TTapp.models import StabilizationThroughPeriods


def do_assign(module, course_type, period, book):
    already_done = ModuleTutorRepartition.objects.filter(
        module=module, course_type=course_type, period=period
    ).exists()
    if already_done:
        return

    assignation_sheet = book["ModuleTutorsAssignation"]
    assign_ok = False
    for assignation_row in range(1, 100):
        if (
            assignation_sheet.cell(row=assignation_row, column=1).value == module.abbrev
            and assignation_sheet.cell(row=assignation_row, column=2).value
            == course_type.name
        ):
            assign_ok = True
            break
    if not assign_ok:
        raise ValueError(
            f"Rien n'est prévu pour assigner {module.abbrev} / {course_type.name}..."
        )
    column = 3
    tutor_username = assignation_sheet.cell(row=assignation_row, column=column).value
    while tutor_username is not None:
        tutor = Tutor.objects.get(username=tutor_username)
        mtr = ModuleTutorRepartition(
            module=module, course_type=course_type, period=period, tutor=tutor
        )
        nb = assignation_sheet.cell(row=assignation_row + 1, column=column).value
        if nb is not None:
            nb = int(nb)
            mtr.courses_nb = nb
        mtr.save()
        column += 1
        tutor_username = assignation_sheet.cell(
            row=assignation_row, column=column
        ).value
    print(f"Assignation done for {module.abbrev} / {course_type.name}!")


@transaction.atomic
def read_planif_scheduling_period(
    department, book, sheet_name, period: SchedulingPeriod, courses_to_stabilize=None
):
    sheet = book[sheet_name]
    training_period = TrainingPeriod.objects.get(name=sheet_name, department=department)
    Course.objects.filter(
        type__department=department,
        period=period,
        module__training_period=training_period,
    ).delete()
    after_type_dependencies = []
    # lookup period column
    short_period_name = period.name.split("-")[0]
    wc = 7
    for wr in [1]:
        while wc < 100:
            wc += 1
            col_period_name = sheet.cell(row=wr, column=wc).value
            if col_period_name is None or col_period_name == "VERIF":
                print(f"Pas de période {short_period_name} en {sheet_name}")
                return
            if short_period_name == col_period_name:
                period_col = wc
                break
    print(f"Période {short_period_name} de {sheet_name} : colonne {period_col}")

    row = 4
    module_col = 1
    nature_col = 3
    duration_col = 4
    prof_col = 5
    room_type_col = 6
    group_col = 7
    sumtotal = 0
    comments = []

    while 1:
        row += 1
        if courses_to_stabilize is not None:
            if row not in courses_to_stabilize:
                courses_to_stabilize[row] = []
        is_total = sheet.cell(row=row, column=group_col).value
        if is_total == "TOTAL":
            # print "Period %g de %s - TOTAL: %g"%(period, feuille,sumtotal)
            break

        cell = sheet.cell(row=row, column=period_col)
        courses_number = cell.value
        if courses_number is None:
            continue

        try:
            room_type_name = sheet.cell(row=row, column=room_type_col).value
            module_abbrev = sheet.cell(row=row, column=module_col).value
            courses_number = float(courses_number)
            # handle dark green lines - Vert fonce
            assert isinstance(room_type_name, str) and room_type_name is not None
            if room_type_name == "Type de Salle":
                nominal = int(courses_number)
                if courses_number != nominal:
                    print(
                        f"Valeur decimale ligne {row} de {sheet_name}, période {col_period_name} : "
                        f" on la met a 1 !"
                    )
                    nominal = 1
                    # le nominal est le nombre de cours par groupe (de TP ou TD)
                if cell.comment:
                    comments = (
                        cell.comment.text.replace(" ", "")
                        .replace("\n", "")
                        .replace(",", ";")
                        .split(";")
                    )
                else:
                    comments = []

                sumtotal += nominal

                continue

            # handle light green lines - Vert clair
            module = Module.objects.get(
                abbrev=module_abbrev, training_period=training_period
            )
            train_prog = module.train_prog
            type_name = sheet.cell(row=row, column=nature_col).value
            tutor_usernames = sheet.cell(row=row, column=prof_col).value
            groups_names = sheet.cell(row=row, column=group_col).value
            course_type = CourseType.objects.get(name=type_name, department=department)
            room_type = RoomType.objects.get(name=room_type_name, department=department)
            duration = dt.timedelta(
                minutes=sheet.cell(row=row, column=duration_col).value
            )
            supp_tutors_usernames = []
            possible_tutors_usernames = []
            if tutor_usernames is None:
                tutor = None
            elif tutor_usernames == "*":
                tutor = None
                do_assign(module, course_type, period, book)
            else:
                assert isinstance(tutor_usernames, str)
                tutor_usernames = tutor_usernames.replace("\xa0", "").replace(" ", "")
                if "|" in tutor_usernames:
                    possible_tutors_usernames = tutor_usernames.split("|")
                    tutor = None
                else:
                    tutors_usernames_list = tutor_usernames.split(";")
                    tutor_username = tutors_usernames_list[0]
                    tutor = Tutor.objects.get(username=tutor_username)
                    supp_tutors_usernames = tutors_usernames_list[1:]

            if cell.comment:
                local_comments = (
                    cell.comment.text.replace("\xa0", "")
                    .replace(" ", "")
                    .replace("\n", "")
                    .replace(",", ";")
                    .split(";")
                )
            else:
                local_comments = []

            all_comments = comments + local_comments

            if isinstance(groups_names, (int, float)):
                groups_names = str(int(groups_names))
            if not groups_names:
                groups_names = []
            else:
                groups_names = (
                    groups_names.replace("\xa0", "")
                    .replace(" ", "")
                    .replace(",", ";")
                    .split(";")
                )
            groups_names_list = [str(g) for g in groups_names]

            groups = list(
                GenericGroup.objects.filter(
                    name__in=groups_names_list, train_prog=train_prog
                )
            )
            if not groups:
                raise ValueError(
                    f"Group(s) do(es) not exist {row}, period {col_period_name} of {sheet_name}\n"
                )

            courses_number = int(courses_number)

            for i in range(courses_number):
                course = Course(
                    tutor=tutor,
                    type=course_type,
                    module=module,
                    period=period,
                    room_type=room_type,
                    duration=duration,
                )
                course.save()
                if courses_to_stabilize is not None:
                    courses_to_stabilize[row].append(course)
                for g in groups:
                    course.groups.add(g)
                course.save()
                if supp_tutors_usernames != []:
                    supp_tutors = Tutor.objects.filter(
                        username__in=supp_tutors_usernames
                    )
                    for sp in supp_tutors:
                        course.supp_tutors.add(sp)
                    course.save()
                if possible_tutors_usernames != []:
                    cpt = CoursePossibleTutors(course=course)
                    cpt.save()
                    for pp in possible_tutors_usernames:
                        t = Tutor.objects.get(username=pp)
                        cpt.possible_tutors.add(t)
                    cpt.save()

                for after_type in [x for x in all_comments if x[0] == "A"]:
                    try:
                        n = int(after_type[1])
                        s = 2
                    except ValueError:
                        n = 1
                        s = 1
                    course_type = after_type[s:]
                    relevant_groups = set()
                    for g in groups:
                        relevant_groups |= (
                            g.ancestor_groups() | {g} | g.descendants_groups()
                        )
                    course_type_queryset = Course.objects.filter(
                        type__name=course_type,
                        module=module,
                        period=period,
                        groups__in=relevant_groups,
                    ).exclude(id=course.id)
                    for relevant_group in relevant_groups:
                        courses_queryset = course_type_queryset.filter(
                            groups=relevant_group
                        )
                        if courses_queryset.exists():
                            after_type_dependencies.append(
                                (course.id, courses_queryset, n, row)
                            )

                if "P" in all_comments:
                    course_additional, _ = CourseAdditional.objects.get_or_create(
                        course=course
                    )
                    course_additional.visio_preference_value = 0
                    course_additional.save()
                elif "DI" in all_comments:
                    course_additional, _ = CourseAdditional.objects.get_or_create(
                        course=course
                    )
                    course_additional.visio_preference_value = 8
                    course_additional.save()
                if "E" in all_comments:
                    course_additional, _ = CourseAdditional.objects.get_or_create(
                        course=course
                    )
                    course_additional.graded = True
                    course_additional.save()
            if "D" in comments or "D" in local_comments and courses_number >= 2:
                relevant_courses = Course.objects.filter(
                    type=course_type, module=module, groups__in=groups, period=period
                )
                for i in range(courses_number // 2):
                    dependency = Dependency(
                        course1=relevant_courses[2 * i],
                        course2=relevant_courses[2 * i + 1],
                        successive=True,
                    )
                    dependency.save()
            if "ND" in comments or "ND" in local_comments and courses_number >= 2:
                relevant_courses = Course.objects.filter(
                    type=course_type, module=module, groups__in=groups, period=period
                )
                for i in range(courses_number - 1):
                    dependency = Dependency(
                        course1=relevant_courses[i],
                        course2=relevant_courses[i + 1],
                        day_gap=1,
                    )
                    dependency.save()

        except Exception as e:
            raise ValueError(
                f"Exception ligne {row}, période {col_period_name} de {sheet_name}: {e} \n"
            ) from e

    # Add after_type dependecies
    for course_id, courses_queryset, n, row in after_type_dependencies:
        course2 = Course.objects.get(id=course_id)
        for course1 in courses_queryset[:n]:
            dependency = Dependency.objects.create(course1=course1, course2=course2)


@transaction.atomic
def extract_training_period(
    department,
    book,
    training_period: TrainingPeriod,
    stabilize_courses=False,
    periods=None,
):
    if stabilize_courses:
        courses_to_stabilize = {}
        print(
            "Courses will be stabilized through scheduling periods for training period",
            training_period,
        )
    else:
        courses_to_stabilize = None
    considered_periods = set(training_period.periods.all())
    if periods is not None:
        considered_periods &= set(periods)
    considered_periods = list(considered_periods)
    considered_periods.sort()
    for period in considered_periods:
        read_planif_scheduling_period(
            department, book, training_period.name, period, courses_to_stabilize
        )

    if stabilize_courses:
        for courses_list in courses_to_stabilize.values():
            if len(courses_list) < 2:
                continue
            stw = StabilizationThroughPeriods.objects.create(department=department)
            for c in courses_list:
                stw.courses.add(c)


@transaction.atomic
def extract_planif(
    department,
    bookname=None,
    stabilize_courses=False,
    scheduling_periods=None,
    training_periods=None,
    assign_colors=True,
):
    """
    Generate the courses from bookname; the school year starts in actual_year
    """
    if bookname is None:
        bookname = os.path.join(
            ds.MEDIA_ROOT,
            "media/configuration/planif_file_" + department.abbrev + ".xlsx",
        )
    book = load_workbook(filename=bookname, data_only=True)
    training_periods = define_training_periods(department, book, training_periods)
    for training_period in training_periods:
        extract_training_period(
            department,
            book,
            training_period,
            stabilize_courses,
            periods=scheduling_periods,
        )
    if assign_colors:
        assign_module_color(department, overwrite=True)


@transaction.atomic
def extract_planif_scheduling_periods(
    scheduling_periods, department, bookname=None, training_periods=None
):
    if bookname is None:
        bookname = os.path.join(
            ds.MEDIA_ROOT,
            "media/configuration/planif_file_" + department.abbrev + ".xlsx",
        )
    book = load_workbook(filename=bookname, data_only=True)
    training_periods = define_training_periods(department, book, training_periods)
    for training_period in training_periods:
        for scheduling_period in scheduling_periods:
            read_planif_scheduling_period(
                department, book, training_period.name, scheduling_period
            )


def define_training_periods(department, book, training_periods):
    if training_periods is None:
        training_periods = TrainingPeriod.objects.filter(department=department)
    ok_training_periods = []
    for training_period in training_periods:
        if training_period.name in book:
            ok_training_periods.append(training_period)
    return ok_training_periods
