#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# This file is part of the FlOpEDT/FlOpScheduler project.
# Copyright (c) 2017
# Authors: Iulian Ober, Paul Renaud-Goud, Pablo Seban, et al.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public
# License along with this program. If not, see
# <http://www.gnu.org/licenses/>.
#
# You can be released from the requirements of the license by purchasing
# a commercial license. Buying such a license is mandatory as soon as
# you develop activities involving the FlOpEDT/FlOpScheduler software
# without disclosing the source code of your own applications.

# This code takes a XLSX database file description and turns it with
# the main function database_description_load_xlsx_file into
# structured Python data, following the data structure described in
# database_description_checker.py.
#

import datetime as dt
import logging
from django.conf import settings as ds

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from base.timing import Day

logger = logging.getLogger(__name__)

PEOPLE_SHEET = "Intervenants"
ROOMS_SHEET = "Salles"
GROUPS_SHEET = "Groupes"
MODULES_SHEET = "Modules"
COURSES_SHEET = "Cours"
SETTINGS_SHEET = "Paramètres"

REASONABLE = 3141  # large enough?


# trivial helper
def cell_name(row, column):
    return f"{get_column_letter(column)}{row}"


#################################################
#                                               #
#   Helper functions to parse various types     #
#   of cells, either individually or by range   #
#                                               #
#################################################


def parse_integer(sheet, row, column):
    try:
        return int(sheet.cell(row=row, column=column).value)
    except ValueError:
        return None


def parse_time(sheet, row, column):
    """
    Helper function to get a time out of a cell
    as a datetime.time object
    (will return None if anything goes wrong)
    """

    val = sheet.cell(row=row, column=column).value
    if isinstance(val, dt.time):
        return val
    if isinstance(val, str):
        try:
            return dt.datetime.strptime(val, "%H:%M").time()
        except ValueError:
            return None
    return None


def parse_date(sheet, row, column):
    """
    Helper function to get a date out of a cell
    as a datetime.date object
    (will return None if anything goes wrong)
    """
    val = sheet.cell(row=row, column=column).value
    if isinstance(val, dt.datetime):
        return val.date()
    if isinstance(val, dt.date):
        return val
    if isinstance(val, str):
        val = " ".split(val, maxsplit=1)[0]
        try:
            return dt.date.fromisoformat(val)
        except ValueError:
            return None
    return None


def parse_time_list_in_line(sheet, row, col_start):
    """
    Parse a line representing a list of times
    (stop at the first empty cell)
    """

    result = []
    col = col_start
    while col < REASONABLE:
        val = parse_time(sheet, row, col)
        if val is None:
            break
        result.append(val)
        col = col + 1

    return result


def parse_string(sheet, row, column):
    """
    Helper function to get a clean string out of a cell
    (will return '' if there's nothing to see)
    """

    val = sheet.cell(row=row, column=column).value
    if val is None:
        val = ""
    val = str(val).strip()

    return val


def parse_string_set_in_line(sheet, row, col_start):
    """
    Parse a line representing a set of strings
    (stop at the first empty cell)
    """

    result = set()
    col = col_start
    while col < REASONABLE:
        val = parse_string(sheet, row, col)
        if val == "":
            break
        result.add(val)
        col = col + 1

    return result


def parse_string_set_dictionary(sheet, row_start, col_start, row_end=REASONABLE):
    """
    Parse a block, turning it into a dictionary of string sets
    The first column gives the keys, the line at the right of the key is
    the associated value
    """
    result = {}
    row = row_start
    while row < row_end:
        name = parse_string(sheet, row, col_start)
        if name == "":
            row = row + 1
            continue
        if name in result:
            name = f":INVALID:DUPLICATE:{cell_name(row, col_start)}"
        result[name] = parse_string_set_in_line(sheet, row, col_start + 1)
        row = row + 1

    return result


################################
#                              #
#   Various helper functions   #
#                              #
################################


def find_marker_cell(sheet, marker, row=1, col=1):
    """
    Helper function to find the marker of a data block
    (Will return either row, col or None, None)
    (With optional parameters to start the search at some position)
    """

    while row < REASONABLE:
        while col < REASONABLE:
            if (
                parse_string(sheet, row, col) == marker
                and sheet.cell(row=row, column=col).font.size > 10
            ):
                return row, col
            col = col + 1
        row = row + 1
        col = 1
    return None, None


def time_from_integer(time: int):
    hours = time // 60
    minutes = time % 60
    return dt.time(hour=hours, minute=minutes)


def strftime_from_time(time: dt.time):
    return time.strftime("%H:%M")


#################################################
#                                               #
#   Parser functions for the different pages    #
#                                               #
#################################################


def parse_rooms(sheet):
    row_groups, col_groups = find_marker_cell(sheet, "Groupes")
    row_cats, col_cats = find_marker_cell(sheet, "Catégories")
    if col_groups != col_cats or row_cats < row_groups:
        logger.warning("The marker cells in sheet %s are misplaced", ROOMS_SHEET)
        return set(), {}, {}

    #
    # parse the groups
    #
    groups = parse_string_set_dictionary(sheet, row_groups + 1, col_groups, row_cats)

    #
    # parse the categories
    #
    categories = parse_string_set_dictionary(sheet, row_cats + 1, col_cats)

    #
    # Build the set of rooms
    #
    rooms = set()
    for lst in groups.values():
        rooms.update(lst)
    for lst in categories.values():
        rooms.update(lst)
    rooms.difference_update(groups.keys())
    rooms.difference_update(categories.keys())

    return rooms, groups, categories


def parse_people(sheet):
    row, col = find_marker_cell(sheet, "Identifiant")
    if row is None:
        return {}

    row = row + 1
    result = {}
    while row < REASONABLE:
        id_ = parse_string(sheet, row, col)
        if id_ == "":
            row = row + 1
            continue
        if id_ in result:
            id_ = f":INVALID:DUPLICATE:{cell_name(row, col)}"
        result[id_] = {
            "last_name": parse_string(sheet, row, col + 1),
            "first_name": parse_string(sheet, row, col + 2),
            "email": parse_string(sheet, row, col + 3),
            "status": parse_string(sheet, row, col + 4),
            "employer": parse_string(sheet, row, col + 5),
        }
        row = row + 1
    return result


def parse_modules(sheet):
    row, col = find_marker_cell(sheet, "Identifiant")
    if row is None:
        logger.warning("The marker cell in sheet %s is missing", MODULES_SHEET)
        return {}

    row = row + 1
    result = {}
    while row < REASONABLE:
        id_ = parse_string(sheet, row, col)
        if id_ == "":
            row = row + 1
            continue
        if id_ in result:
            id_ = f":INVALID:DUPLICATE:{cell_name(row, col)}"
        result[id_] = {
            "short": parse_string(sheet, row, col + 1),
            "PPN": parse_string(sheet, row, col + 2),
            "name": parse_string(sheet, row, col + 3),
            "promotion": parse_string(sheet, row, col + 4),
            "period": parse_string(sheet, row, col + 5),
            "responsable": parse_string(sheet, row, col + 6),
        }
        row = row + 1
    return result


def parse_courses(sheet):
    row_type, col_type = find_marker_cell(sheet, "Type de cours")
    if row_type is None:
        logger.warning("The marker cell 'Type' in sheet %s is missing", COURSES_SHEET)
        return {}, {}

    row_constraint, col_constraint = find_marker_cell(sheet, "Durée de cours")
    if row_constraint is None:
        logger.warning(
            "The marker 'Duration' cell in sheet %s is missing", COURSES_SHEET
        )
        return {}

    row_type += 1
    course_types = {}
    while row_type < row_constraint:
        id_ = parse_string(sheet, row_type, col_type)
        if id_ == "":
            row_type = row_type + 1
            continue
        if id_ in course_types:
            id_ = f":INVALID:DUPLICATE:{cell_name(row_type, col_type)}"
        course_types[id_] = {
            "graded": parse_string(sheet, row_type, col_type + 1),
            "group_types": set(parse_string_set_in_line(sheet, row_type, col_type + 2)),
        }
        row_type += 1

    row_constraint += 1
    course_start_time_constraints = {}
    while row_constraint < REASONABLE:
        duration_str = parse_string(sheet, row_constraint, col_constraint)
        if duration_str == "":
            row_constraint = row_constraint + 1
            continue
        if duration_str in course_start_time_constraints:
            duration_str = (
                f":INVALID:DUPLICATE:{cell_name(row_constraint, col_constraint)}"
            )
        course_start_time_constraints[duration_str] = {
            "start_times": set(
                parse_time_list_in_line(sheet, row_constraint, col_constraint + 1)
            )
        }
        row_constraint += 1

    return course_types, course_start_time_constraints


def parse_settings(sheet):
    result = {}
    row, col = find_marker_cell(sheet, "Jalon")
    if row is None:
        logger.warning("The 'Jalon' cell in sheet %s is missing", SETTINGS_SHEET)
        result["day_start_time"] = None
        result["day_end_time"] = None
        result["morning_end_time"] = None
        result["afternoon_start_time"] = None
    else:
        val = parse_time(sheet, row + 1, col + 1)
        result["day_start_time"] = val
        val = parse_time(sheet, row + 2, col + 1)
        result["day_end_time"] = val
        val = parse_time(sheet, row + 3, col + 1)
        result["morning_end_time"] = val
        val = parse_time(sheet, row + 4, col + 1)
        result["afternoon_start_time"] = val

    days = []
    row, col = find_marker_cell(sheet, "Jours ouvrables")
    if row is not None:
        for index, choice in enumerate(Day.CHOICES):
            if parse_string(sheet, row + 2, col + index) != "":
                days.append(choice[0])
    result["days"] = days

    row, col = find_marker_cell(sheet, "Modes")
    visio_mode_str = parse_string(sheet, row + 1, col)
    cosmo_mode_str = parse_integer(sheet, row + 2, col)
    scheduling_mode_str = parse_string(sheet, row + 3, col)

    visio_mode = visio_mode_str == "Visio"
    cosmo_mode = 0
    if cosmo_mode_str == "Educatif":
        cosmo_mode = 0
    elif cosmo_mode_str == "Coop.(Poste)":
        cosmo_mode = 1
    elif cosmo_mode_str == "Coop. (Salarié)":
        cosmo_mode = 2
    scheduling_mode = "w"
    if scheduling_mode_str == "Par semaine":
        scheduling_mode = "w"
    elif scheduling_mode_str == "Par jour":
        scheduling_mode = "d"
    elif scheduling_mode_str == "Par mois":
        scheduling_mode = "m"
    elif scheduling_mode_str == "Par an":
        scheduling_mode = "y"
    elif scheduling_mode_str == "Custom":
        scheduling_mode = "c"

    result["mode"] = {
        "visio": visio_mode,
        "cosmo": cosmo_mode,
        "scheduling_mode": scheduling_mode,
    }

    periods = {}
    row, col = find_marker_cell(sheet, "Périodes de cours")
    if row is not None:
        row = row + 2
        while row < REASONABLE:
            id_ = parse_string(sheet, row, col)
            if id_ == "":
                row = row + 1
                continue
            if id_ in periods:
                id_ = f":INVALID:DUPLICATE:{cell_name(row, col)}"
            start_date = parse_date(sheet, row, col + 1)
            end_date = parse_date(sheet, row, col + 2)
            periods[id_] = (start_date, end_date)
            row = row + 1
    result["training_periods"] = periods

    return result


def parse_groups(sheet):
    row_prom, col_prom = find_marker_cell(sheet, "Identifiant")
    if row_prom is None:
        return {}, set(), {}, {}

    row_nat, col_nat = find_marker_cell(sheet, "Identifiant", row_prom + 1)
    if row_nat is None:
        return {}, set(), {}, {}

    row_grp, col_grp = find_marker_cell(sheet, "Identifiant", row_nat + 1)
    if row_grp is None:
        return {}, set(), {}, {}

    row_trans_grp, col_trans_grp = find_marker_cell(sheet, "Identifiant", row_grp + 1)
    if row_trans_grp is None:
        return {}, set(), {}, {}

    promotions = {}
    row = row_prom + 1
    while row < row_nat:  # should stop before
        id_ = parse_string(sheet, row, col_prom)
        if id_ == "":
            break
        if id_ in promotions:
            id_ = f":INVALID:DUPLICATE:{cell_name(row, col_prom)}"
        promotions[id_] = parse_string(sheet, row, col_prom + 1)
        row = row + 1

    group_types = set()
    row = row_nat + 1
    while row < row_grp:  # should stop before
        id_ = parse_string(sheet, row, col_nat)
        if id_ == "":
            break
        if id_ in group_types:
            id_ = f":INVALID:DUPLICATE:{cell_name(row, col_prom)}"
        group_types.add(id_)
        row = row + 1

    structural_groups = {}
    row = row_grp + 1
    while row < row_trans_grp - 4:
        id_ = parse_string(sheet, row, col_grp)
        if id_ == "":
            row += 1
            continue
        if id_ in structural_groups:
            id_ = f":INVALID:DUPLICATE:{cell_name(row, col_prom)}"
        promotion = parse_string(sheet, row, col_grp + 1)
        group_type = parse_string(sheet, row, col_grp + 2)
        parent_ = parse_string(sheet, row, col_grp + 3)
        if parent_ == "":
            parent = set()
        else:
            parent = {parent_}
        structural_groups[promotion, id_] = {"group_type": group_type, "parent": parent}
        row = row + 1

    transversal_groups = {}
    row = row_trans_grp + 1
    while row < REASONABLE:
        id_ = parse_string(sheet, row, col_grp)
        if id_ == "":
            row += 1
            continue
        if id_ in set(structural_groups) | set(transversal_groups):
            id_ = f":INVALID:DUPLICATE:{cell_name(row, col_prom)}"
        promotion = parse_string(sheet, row, col_trans_grp + 1)
        _, col_trans = find_marker_cell(sheet, "Groupes structuraux en conflit")
        _, col_par = find_marker_cell(sheet, "Groupes transversaux parallèles")
        transversal_to = parse_string_set_in_line(sheet, row, col_trans)
        parallel_to = parse_string_set_in_line(sheet, row, col_par)

        transversal_groups[promotion, id_] = {
            "transversal_to": transversal_to,
            "parallel_to": parallel_to,
        }
        row = row + 1

    return promotions, group_types, structural_groups, transversal_groups


#################################################
#                                               #
#           Main parsing function               #
#                                               #
#################################################


def database_description_load_xlsx_file(filename="file_essai.xlsx"):
    try:
        wb = load_workbook(filename, data_only=True)

        sheet = wb[ROOMS_SHEET]
        if not sheet:
            logger.warning("Sheet %s doesn't exist", ROOMS_SHEET)
            return None

        rooms, room_groups, room_categories = parse_rooms(sheet)

        sheet = wb[PEOPLE_SHEET]
        if not sheet:
            logger.warning("Sheet %s doesn't exist", PEOPLE_SHEET)
            return None

        people = parse_people(sheet)

        sheet = wb[MODULES_SHEET]
        if not sheet:
            logger.warning("Sheet %s doesn't exist", MODULES_SHEET)
            return None

        modules = parse_modules(sheet)

        sheet = wb[COURSES_SHEET]
        if not sheet:
            logger.warning("Sheet %s doesn't exist", COURSES_SHEET)
            return None

        course_types, course_start_time_constraints = parse_courses(sheet)

        sheet = wb[SETTINGS_SHEET]
        if not sheet:
            logger.warning("Sheet %s doesn't exist", SETTINGS_SHEET)
            return None

        settings = parse_settings(sheet)

        sheet = wb[GROUPS_SHEET]
        if not sheet:
            logger.warning("Sheet %s doesn't exist", GROUPS_SHEET)
            return None

        promotions, group_types, structural_groups, transversal_groups = parse_groups(
            sheet
        )

        return {
            "rooms": rooms,
            "room_groups": room_groups,
            "room_categories": room_categories,
            "people": people,
            "modules": modules,
            "course_types": course_types,
            "course_start_time_constraints": course_start_time_constraints,
            "settings": settings,
            "promotions": promotions,
            "group_types": group_types,
            "groups": structural_groups,
            "transversal_groups": transversal_groups,
        }
    except FileNotFoundError as ex:
        logger.warning("Database file couldn't be opened: %s", ex)
        return None


def database_description_save_xlsx_file(filename, database_dict):
    wb = load_workbook(f"{ds.MEDIA_ROOT}/empty_database_file.xlsx")

    sheet = wb[SETTINGS_SHEET]
    row, col = find_marker_cell(sheet, "Jalon")
    sheet.cell(
        row=row + 1,
        column=col + 1,
        value=time_from_integer(database_dict["settings"]["day_start_time"]),
    )
    sheet.cell(
        row=row + 2,
        column=col + 1,
        value=time_from_integer(database_dict["settings"]["day_finish_time"]),
    )
    sheet.cell(
        row=row + 3,
        column=col + 1,
        value=time_from_integer(database_dict["settings"]["lunch_break_start_time"]),
    )
    sheet.cell(
        row=row + 4,
        column=col + 1,
        value=time_from_integer(database_dict["settings"]["lunch_break_finish_time"]),
    )

    row, col = find_marker_cell(sheet, "Granularité")
    sheet.cell(
        row=row,
        column=col + 1,
        value=database_dict["settings"]["default_preference_duration"],
    )

    row, col = find_marker_cell(sheet, "Jours ouvrables")
    days = database_dict["settings"]["days"]
    cols = {"m": 0, "tu": 1, "w": 2, "th": 3, "f": 4, "sa": 5, "su": 6}
    for day, delta in cols.items():
        if day in days:
            sheet.cell(row=row + 2, column=col + delta, value="X")
        else:
            sheet.cell(row=row + 2, column=col + delta, value=None)

    row, col = find_marker_cell(sheet, "Périodes")
    row = row + 1
    for id_, (start, finish) in database_dict["settings"]["periods"].items():
        row = row + 1
        sheet.cell(row=row, column=col, value=id_)
        sheet.cell(row=row, column=col + 1, value=start)
        sheet.cell(row=row, column=col + 2, value=finish)

    sheet = wb[PEOPLE_SHEET]
    row, col = find_marker_cell(sheet, "Identifiant")
    for id_, data in database_dict["people"].items():
        row = row + 1
        sheet.cell(row=row, column=col, value=id_)
        sheet.cell(row=row, column=col + 1, value=data["last_name"])
        sheet.cell(row=row, column=col + 2, value=data["first_name"])
        sheet.cell(row=row, column=col + 3, value=data["email"])
        sheet.cell(row=row, column=col + 4, value=data["status"])
        sheet.cell(row=row, column=col + 5, value=data["employer"])

    sheet = wb[ROOMS_SHEET]

    row, col_start = find_marker_cell(sheet, "Groupes")
    for id_, rooms in database_dict["room_groups"].items():
        row = row + 1
        col = col_start
        sheet.cell(row=row, column=col, value=id_)
        for room in rooms:
            col = col + 1
            sheet.cell(row=row, column=col, value=room)

    row, col_start = find_marker_cell(sheet, "Catégories")
    for id_, rooms in database_dict["room_categories"].items():
        row = row + 1
        col = col_start
        sheet.cell(row=row, column=col, value=id_)
        for room in rooms:
            col = col + 1
            sheet.cell(row=row, column=col, value=room)

    sheet = wb[GROUPS_SHEET]

    row, col = find_marker_cell(sheet, "Identifiant")
    for id_, name in database_dict["promotions"].items():
        row = row + 1
        sheet.cell(row=row, column=col, value=id_)
        sheet.cell(row=row, column=col + 1, value=name)

    row, col = find_marker_cell(sheet, "Identifiant", row)
    for id_ in database_dict["group_types"]:
        row = row + 1
        sheet.cell(row=row, column=col, value=id_)

    row, col = find_marker_cell(sheet, "Identifiant", row)
    for id_, data in database_dict["groups"].items():
        row = row + 1
        sheet.cell(row=row, column=col, value=id_)
        sheet.cell(row=row, column=col + 1, value=data["promotion"])
        sheet.cell(row=row, column=col + 2, value=data["group_type"])
        for parent in data["parent"]:
            sheet.cell(row=row, column=col + 3, value=parent)

    sheet = wb[MODULES_SHEET]
    row, col = find_marker_cell(sheet, "Identifiant")
    for id_, data in database_dict["modules"].items():
        row = row + 1
        sheet.cell(row=row, column=col, value=id_)
        sheet.cell(row=row, column=col + 1, value=data["short"])
        sheet.cell(row=row, column=col + 2, value=data["PPN"])
        sheet.cell(row=row, column=col + 3, value=data["name"])
        sheet.cell(row=row, column=col + 4, value=data["promotion"])
        sheet.cell(row=row, column=col + 5, value=data["period"])
        sheet.cell(row=row, column=col + 6, value=data["responsable"])

    sheet = wb[COURSES_SHEET]
    row, col_start = find_marker_cell(sheet, "Type")
    for id_, data in database_dict["courses"].items():
        row = row + 1
        col = col_start
        sheet.cell(row=row, column=col, value=id_)
        sheet.cell(row=row, column=col + 1, value=data["duration"])
        col = col_start + 1
        for group_type in data["group_types"]:
            col = col + 1
            sheet.cell(row=row, column=col, value=group_type)
        row = row + 1
        col = col_start + 1
        start_times = list(data["start_times"])
        start_times.sort()
        for start_time in start_times:
            col = col + 1
            sheet.cell(row=row, column=col, value=time_from_integer(start_time))

    wb.save(filename)
