#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# This file is part of the FlOpEDT/FlOpScheduler project.
# Copyright (c) 2017
# Authors: Iulian Ober, Paul Renaud-Goud, Pablo Seban, et al.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public
# License along with this program. If not, see
# <http://www.gnu.org/licenses/>.
#
# You can be released from the requirements of the license by purchasing
# a commercial license. Buying such a license is mandatory as soon as
# you develop activities involving the FlOpEDT/FlOpScheduler software
# without disclosing the source code of your own applications.

# This code takes a XLSX database file description and turns it with
# the main function database_description_load_xlsx_file into
# structured Python data, following the data structure described in
# database_description_checker.py.
#

import logging
import os
from copy import copy
from django.conf import settings as ds

from openpyxl import load_workbook

from base.models import (
    CourseStartTimeConstraint,
    CourseType,
    GroupType,
    Module,
    Room,
    RoomType,
    StructuralGroup,
    TrainingPeriod,
    TrainingProgramme,
    TransversalGroup,
)
from configuration.database_description_xlsx import (
    COURSES_SHEET,
    find_marker_cell,
    GROUPS_SHEET,
    MODULES_SHEET,
    PEOPLE_SHEET,
    ROOMS_SHEET,
    SETTINGS_SHEET,
    strftime_from_time,
)
from people.models import Tutor

logger = logging.getLogger(__name__)
#################################################
#                                               #
#   Filler functions for the different pages    #
#                                               #
#################################################
MAX_ROW = 200
MAX_COL = 20


def insert_missing_rows(sheet, row_rank, data, existing_rows_number):
    merged_cells = {r.bounds[1]: [] for r in sheet.merged_cells.ranges}
    for r in sheet.merged_cells.ranges:
        merged_cells[r.bounds[1]].append((r.bounds[0], r.bounds[2]))
    # merges as dict {row : [(start_col, end_col),(st,end)], ... }
    missing_rows_number = len(data) - existing_rows_number

    if missing_rows_number > 0:
        # create copy of lines from title_row_rang +1 to MAX_ROW some row downer
        for row in range(MAX_ROW, row_rank, -1):
            new_row = row + missing_rows_number
            if new_row in merged_cells:
                for start_col, end_col in merged_cells[new_row]:
                    sheet.unmerge_cells(
                        start_row=new_row,
                        start_column=start_col,
                        end_row=new_row,
                        end_column=end_col,
                    )
            for col in range(1, MAX_COL):
                cell = sheet.cell(row, col)
                new_cell = sheet.cell(new_row, col)
                new_cell.value = cell.value
                new_cell._style = copy(cell._style)  # pylint: disable=protected-access
            if row in merged_cells:
                for start_col, end_col in merged_cells[row]:
                    sheet.merge_cells(
                        start_row=new_row,
                        start_column=start_col,
                        end_row=new_row,
                        end_column=end_col,
                    )

        for row in range(row_rank + 1, row_rank + 1 + missing_rows_number):
            if row in merged_cells:
                for start_col, end_col in merged_cells[row]:
                    sheet.unmerge_cells(
                        start_row=row,
                        start_column=start_col,
                        end_row=row,
                        end_column=end_col,
                    )
            for col in range(1, MAX_COL):
                cell = sheet.cell(row_rank + 1, col)
                new_cell = sheet.cell(row, col)
                new_cell._style = copy(cell._style)  # pylint: disable=protected-access
                new_cell.value = cell.value
            if row in merged_cells:
                for start_col, end_col in merged_cells[row]:
                    sheet.merge_cells(
                        start_row=row,
                        start_column=start_col,
                        end_row=row,
                        end_column=end_col,
                    )


def dict_from_dept_database(department):
    database_dict = {}
    database_dict["settings"] = {}
    settings = department.timegeneralsettings
    database_dict["settings"]["day_start_time"] = settings.day_start_time
    database_dict["settings"]["day_end_time"] = settings.day_end_time
    database_dict["settings"]["morning_end_time"] = settings.morning_end_time
    database_dict["settings"]["afternoon_start_time"] = settings.afternoon_start_time


def make_filled_database_file(department, filename=None):
    wb = load_workbook(
        os.path.join(os.path.dirname(__file__), "xls/empty_database_file.xlsx")
    )
    if filename is None:
        filename = os.path.join(
            ds.CONF_XLS_DIR, f"database_file_{department.abbrev}.xlsx"
        )
        print(filename)
    sheet = wb[SETTINGS_SHEET]
    row, col = find_marker_cell(sheet, "Jalon")
    sheet.cell(
        row=row + 1,
        column=col + 1,
        value=strftime_from_time(department.timegeneralsettings.day_start_time),
    )
    sheet.cell(
        row=row + 2,
        column=col + 1,
        value=strftime_from_time(department.timegeneralsettings.day_end_time),
    )
    sheet.cell(
        row=row + 3,
        column=col + 1,
        value=strftime_from_time(department.timegeneralsettings.morning_end_time),
    )
    sheet.cell(
        row=row + 4,
        column=col + 1,
        value=strftime_from_time(department.timegeneralsettings.afternoon_start_time),
    )

    row, col = find_marker_cell(sheet, "Modes")
    mode = department.mode
    if mode.visio:
        sheet.cell(row=row + 1, column=col, value="Visio")
    else:
        sheet.cell(row=row + 1, column=col, value="No Visio")
    if mode.cosmo == 0:
        sheet.cell(row=row + 2, column=col, value="Educatif")
    elif mode.cosmo == 1:
        sheet.cell(row=row + 2, column=col, value="Coop.(Poste)")
    else:
        sheet.cell(row=row + 2, column=col, value="Coop. (Salarié)")
    if mode.scheduling_mode == "w":
        sheet.cell(row=row + 3, column=col, value="Par semaine")
    elif mode.scheduling_mode == "d":
        sheet.cell(row=row + 3, column=col, value="Par jour")
    elif mode.scheduling_mode == "m":
        sheet.cell(row=row + 3, column=col, value="Par mois")
    elif mode.scheduling_mode == "y":
        sheet.cell(row=row + 3, column=col, value="Par an")
    else:
        sheet.cell(row=row + 3, column=col, value="Custom")

    row, col = find_marker_cell(sheet, "Jours ouvrables")
    days = department.timegeneralsettings.days
    cols = {"m": 0, "tu": 1, "w": 2, "th": 3, "f": 4, "sa": 5, "su": 6}
    for day, delta in cols.items():
        if day in days:
            sheet.cell(row=row + 2, column=col + delta, value="X")
        else:
            sheet.cell(row=row + 2, column=col + delta, value=None)

    row, training_period_col = find_marker_cell(sheet, "Périodes de cours")
    row = row + 1
    for training_period in TrainingPeriod.objects.filter(department=department):
        row = row + 1
        sheet.cell(row=row, column=training_period_col, value=training_period.name)
        col = training_period_col
        scheduling_periods = list(training_period.periods.all())
        scheduling_periods.sort(key=lambda x: x.start_date)
        if len(scheduling_periods) > 0:
            first = scheduling_periods[0]
            last = scheduling_periods[-1]
            sheet.cell(row=row, column=col + 1, value=first.start_date)
            sheet.cell(row=row, column=col + 2, value=last.end_date)

    sheet = wb[PEOPLE_SHEET]
    row, col = find_marker_cell(sheet, "Identifiant")
    for tutor in Tutor.objects.filter(departments=department):
        row = row + 1
        sheet.cell(row=row, column=col, value=tutor.username)
        sheet.cell(row=row, column=col + 1, value=tutor.last_name)
        sheet.cell(row=row, column=col + 2, value=tutor.first_name)
        sheet.cell(row=row, column=col + 3, value=tutor.email)

        sheet.cell(row=row, column=col + 4, value=tutor.status)
        if hasattr(tutor, "supplystaff"):
            sheet.cell(row=row, column=col + 5, value=tutor.supplystaff.employer)

    sheet = wb[ROOMS_SHEET]

    row, col_start = find_marker_cell(sheet, "Groupes")
    room_groups = {
        r for r in Room.objects.filter(departments=department) if not r.is_basic
    }
    insert_missing_rows(sheet, row, room_groups, 9)

    for room_group in room_groups:
        row = row + 1
        col = col_start
        sheet.cell(row=row, column=col, value=room_group.name)
        for basic_room in room_group.basic_rooms():
            col = col + 1
            sheet.cell(row=row, column=col, value=basic_room.name)

    row, col_start = find_marker_cell(sheet, "Catégories")
    room_types = RoomType.objects.filter(department=department)
    insert_missing_rows(sheet, row, room_types, 13)
    for room_type in room_types:
        row = row + 1
        col = col_start
        sheet.cell(row=row, column=col, value=room_type.name)
        for room in room_type.members.all():
            col = col + 1
            sheet.cell(row=row, column=col, value=room.name)

    sheet = wb[GROUPS_SHEET]
    train_progs = TrainingProgramme.objects.filter(department=department)
    row, col = find_marker_cell(sheet, "Identifiant")
    # if we have too many promotions, avoid to destroy the marker below!
    insert_missing_rows(sheet, row, train_progs, 7)
    for tp in train_progs:
        row = row + 1
        sheet.cell(row=row, column=col, value=tp.abbrev)
        sheet.cell(row=row, column=col + 1, value=tp.name)

    wb.save(filename)

    row, col = find_marker_cell(sheet, "Identifiant", row)
    group_types = GroupType.objects.filter(department=department)
    # if we have too many group types, avoid to destroy the marker below!
    insert_missing_rows(sheet, row, group_types, 5)
    for gt in group_types:
        row = row + 1
        sheet.cell(row=row, column=col, value=gt.name)

    row, col = find_marker_cell(sheet, "Identifiant", row)
    structural_groups = StructuralGroup.objects.filter(train_prog__in=train_progs)
    insert_missing_rows(sheet, row, structural_groups, 18)
    for gp in structural_groups:
        row = row + 1
        sheet.cell(row=row, column=col, value=gp.name)
        sheet.cell(row=row, column=col + 1, value=gp.train_prog.abbrev)
        sheet.cell(row=row, column=col + 2, value=gp.type.name)
        if gp.parent_groups.count() == 1:
            sheet.cell(row=row, column=col + 3, value=gp.parent_groups.first().name)

    row, col = find_marker_cell(sheet, "Identifiant", row)
    transversal_groups = TransversalGroup.objects.filter(train_prog__in=train_progs)
    insert_missing_rows(sheet, row, transversal_groups, 18)
    for gp in transversal_groups:
        conflicting_groups = gp.conflicting_groups
        parallel_groups = gp.parallel_groups
        nb_conflicts = conflicting_groups.count()
        if nb_conflicts > 7:
            raise ValueError(f"Trop de groupes en conflit avec {gp.name}")
        row = row + 1
        sheet.cell(row=row, column=col, value=gp.name)
        sheet.cell(row=row, column=col + 1, value=gp.train_prog.abbrev)
        col += 2
        for i, confl_group in enumerate(conflicting_groups.all()):
            sheet.cell(row=row, column=col + 2 + i, value=confl_group.name)
        for i, parallel_group in enumerate(parallel_groups.all()):
            sheet.cell(row=row, column=col + 9 + i, value=parallel_group.name)

    sheet = wb[MODULES_SHEET]
    row, col = find_marker_cell(sheet, "Identifiant")
    modules = Module.objects.filter(train_prog__in=train_progs)
    insert_missing_rows(sheet, row, modules, 27)
    for module in modules:
        row = row + 1
        sheet.cell(row=row, column=col, value=module.abbrev)
        sheet.cell(row=row, column=col + 1, value=module.abbrev)
        sheet.cell(row=row, column=col + 2, value=module.ppn)
        sheet.cell(row=row, column=col + 3, value=module.name)
        sheet.cell(row=row, column=col + 4, value=module.train_prog.abbrev)
        sheet.cell(row=row, column=col + 5, value=module.training_period.name)
        if module.head is not None:
            sheet.cell(row=row, column=col + 6, value=module.head.username)

    sheet = wb[COURSES_SHEET]
    row, col_start = find_marker_cell(sheet, "Type de cours")
    course_types = CourseType.objects.filter(department=department)
    insert_missing_rows(sheet, row, list(course_types), 10)
    for course_type in course_types:
        row = row + 1
        col = col_start
        sheet.cell(row=row, column=col, value=course_type.name)
        if course_type.graded:
            sheet.cell(row=row, column=col + 1, value="Oui")
        else:
            sheet.cell(row=row, column=col + 1, value="Non")
        col = col_start + 1
        for group_type in course_type.group_types.all():
            col = col + 1
            sheet.cell(row=row, column=col, value=group_type.name)

    row, col_start = find_marker_cell(sheet, "Durée de cours")
    course_start_time_constraints = CourseStartTimeConstraint.objects.filter(
        department=department
    )
    insert_missing_rows(sheet, row, list(course_start_time_constraints), 12)
    for course_start_time_constraint in course_start_time_constraints:
        row = row + 1
        col = col_start
        sheet.cell(
            row=row,
            column=col,
            value=course_start_time_constraint.duration.total_seconds() // 60,
        )
        allowed_start_times = course_start_time_constraint.allowed_start_times
        allowed_start_times.sort()
        for start_time in allowed_start_times:
            col = col + 1
            sheet.cell(row=row, column=col, value=start_time)

    wb.save(filename)
